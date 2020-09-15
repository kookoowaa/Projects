#///////////////////////////////////////////////////////////////////////////////////
# File Name......: crawling.py
# Auther.........: Pablo Park
# Written date...: 11 Aug. 2020
# Description....: Crawling sales summary from web dashboard using python + selenium
# Dependancies...: python 3.8.3
#                  selenium 3.141.0
#                  openpyxl 3.0.4
#///////////////////////////////////////////////////////////////////////////////////

from selenium import webdriver
import pandas as pd
import numpy as np
import re
import time
import datetime
import os
from openpyxl import Workbook


# ChromeDriver 파일 위치 변경 필요
driver = webdriver.Chrome(r"C:\Users\Admin\Google Drive\KyoChon Sales Data\chromedriver.exe")

driver.get('http://kyochon.revelup.com')
time.sleep(5)
revel_id = "gyutae"
driver.find_element_by_xpath('//*[@id="auth0-lock-container-1"]/div/div[2]/form/div/div/div/div[2]/div[2]/span/div/div/div/div/div/div/div/div/div/div/div[1]/div/input').send_keys(revel_id)
print()
passwd = "Teddygom850212!"
print()
driver.find_element_by_xpath('//*[@id="auth0-lock-container-1"]/div/div[2]/form/div/div/div/div[2]/div[2]/span/div/div/div/div/div/div/div/div/div/div/div[2]/div[1]/div/input').send_keys(passwd)
driver.find_element_by_xpath('//*[@id="auth0-lock-container-1"]/div/div[2]/form/div/div/div/button').click()
time.sleep(5)
driver.find_element_by_xpath('//*[@id="navigation"]/div/ul/li[2]/a/span').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="table_data_place"]/div[1]/div[1]/ul/li[1]/ul/li[1]/div/div[1]').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="table_data_place"]/div[1]/div[1]/ul/li[1]/ul/li[1]/ul/li[1]/div/div[1]').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="table_data_place"]/div[1]/div[1]/ul/li[1]/ul/li[1]/ul/li[1]/ul/li[1]/div/div[1]').click()
time.sleep(2)

# sales 추출 관련 function
def sto_summary():
    # Stores
    sto = driver.find_element_by_xpath('//*[@id="header"]/div[2]/div[5]/div[2]/span').text
    now = datetime.datetime.now()
    # Dates
    now_date = str(now.year) + '-' + ('0' + str(now.month))[-2:] + '-' + ('0' + str(now.day))[-2:]
    # Time
    now_hour = str(now.hour)
    # Net sales
    acc_sales = driver.find_element_by_xpath('//*[@id="table_data_place"]/div[1]/div[1]/ul/li[1]/ul/li[1]/ul/li[1]/ul/li[1]//ul/li[1]/div/div[2]').text
    # Net transaction
    acc_transaction = driver.find_element_by_xpath('//*[@id="table_data_place"]/div[1]/div[2]/ul/li[1]/ul/li[1]/div/div[2]').text
    # avg check (calculated)]
    try:
        avg_check = float(acc_sales)/float(acc_transaction)
    except ZeroDivisionError:
        avg_check = 0
        
    table1 = pd.DataFrame([sto, now_date, now_hour, acc_sales, acc_transaction, avg_check], 
                          index = ['Store', 'Date', 'Hour', 'Sales', 'Transaction','avg check'],
                          columns = ['Summary'])
    table1 = table1.reset_index()
    return table1

# Range Mix 추출 관련 function
def prod_mix():
    tbl_len = len(driver.find_elements_by_class_name('last-matrix-element'))
    table2 = {}
    for i in range(tbl_len):
        table2[driver.find_elements_by_class_name('product-description-container')[i].text] = {'Qty': driver.find_elements_by_class_name('n_items')[i+1].text, 
                                                                                              'Sales': driver.find_elements_by_class_name('price')[i+1].text}
    table2 = pd.DataFrame(table2).T.reset_index()
    table2 = table2.iloc[:len(table2)-2]
    return table2


# 스토어 별로 Loop 돌면서 함수 실행
driver.find_element_by_xpath('//*[@id="header"]/div[2]/div[5]/div[2]/span').click()
time.sleep(5)
driver.find_element_by_xpath('//*[@id="establishments-tree"]/div/div[2]/div[2]/div[2]/span[1]').click()
n_sto = len(driver.find_elements_by_class_name('fancytree-title.disabled'))

summ = []
#range_mix = []

for i in range(n_sto):
    # 중간의 time.sleep() 함수는 필요에 따라 조정 가능
    # 한국에서 테스트 시(접속 시) 페이지 로딩까지 시간이 걸려 에러가 발생함을 확인하여 여유롭게 sleep time 지정
    if i!=0:
        driver.find_element_by_xpath('//*[@id="header"]/div[2]/div[5]/div[2]/span').click()
        time.sleep(5)
        driver.find_element_by_xpath('//*[@id="establishments-tree"]/div/div[2]/div[2]/div[2]/span[1]').click()
    time.sleep(5)
    driver.find_elements_by_class_name('fancytree-title.disabled')[i].click()
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="navigation"]/div/ul/li[2]/a/span').click()
    time.sleep(5)
    summ.append(sto_summary())
    #driver.find_element_by_xpath('//*[@id="report_sub_nav"]/li[4]/a').click()
    #time.sleep(5)
    #range_mix.append(prod_mix())
    
    print('{} finished'.format(driver.find_element_by_xpath('//*[@id="header"]/div[2]/div[5]/div[2]/span').text))


sales_summ = pd.DataFrame()
for i in range(len(summ)):
    sales_summ = pd.concat([sales_summ, pd.DataFrame([summ[i].iloc[0,1],summ[i].iloc[3,1]]).T])


# excel에 저장
wb = Workbook()

ws = wb.create_sheet('summary')
ws.cell(1,1, value = summ[0].iloc[1,1])
nrow, ncol = sales_summ.shape
for r in range(nrow):
    for c in range(ncol):
        ws.cell(r+3,c+1, value = sales_summ.iloc[r,c])
        
for i in range(len(summ)):
    ws = wb.create_sheet()
    nrow, ncol = summ[i].shape
    for r in range(nrow):
        for c in range(ncol):
            ws.cell(r+1,c+1, value = summ[i].iloc[r,c])
    ws.title = ws.cell(1,2).value
    
    #nrow, ncol = range_mix[i].shape
    #for r in range(nrow):
    #    for c in range(ncol):
    #        ws.cell(r+1,c+4, value = range_mix[i].iloc[r,c])


wb.remove(wb['Sheet'])

now = datetime.datetime.now()
fname = "Closing {y}{m}{d}{h}.xlsx".format(y=now.year, m=('0' + str(now.month))[-2:], d=('0' + str(now.day))[-2:],h=now.hour)
wb.save(fname)